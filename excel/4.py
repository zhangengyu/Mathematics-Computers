import openpyxl

# 文件路径
file1 = r"C:\Users\詹耿羽\Desktop\新建 XLSX 工作表.xlsx"
file2 = r"C:\Users\詹耿羽\Desktop\新建 XLSX 工作表 (3).xlsx"

# 检查规则：姓名为 x1 时，钱必须为 12
target_name = "x1"
target_money = 12

def check_excel(file_path, target_name, target_money):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # 获取活动工作表

    # 遍历每一行，从第二行开始（假设第一行为表头）
    invalid_rows = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        name_cell = row[0]  # 假设姓名在第一列（A列）
        money_cell = row[1]  # 假设钱在第二列（B列）
        if name_cell == target_name and money_cell != target_money:
            invalid_rows.append(row)  # 记录不符合标准的行

    wb.close()
    return invalid_rows

# 检查两个文件
invalid_rows_file1 = check_excel(file1, target_name, target_money)
invalid_rows_file2 = check_excel(file2, target_name, target_money)

# 输出结果
if invalid_rows_file1:
    print(f"文件 {file1} 中不符合标准的行：")
    for row in invalid_rows_file1:
        print(row)
else:
    print(f"文件 {file1} 中所有数据均符合标准。")

if invalid_rows_file2:
    print(f"文件 {file2} 中不符合标准的行：")
    for row in invalid_rows_file2:
        print(row)
else:
    print(f"文件 {file2} 中所有数据均符合标准。")