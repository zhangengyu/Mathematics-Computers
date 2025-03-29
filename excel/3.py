import openpyxl

# 文件路径
file1 = r"C:\Users\詹耿羽\Desktop\新建 XLSX 工作表.xlsx"
file2 = r"C:\Users\詹耿羽\Desktop\新建 XLSX 工作表 (3).xlsx"

# 输出文件路径
output_file1 = r"C:\Users\詹耿羽\Desktop\过滤后的新建 XLSX 工作表.xlsx"
output_file2 = r"C:\Users\詹耿羽\Desktop\过滤后的新建 XLSX 工作表 (3).xlsx"

# 有效的姓名列表
valid_names = ['x1', 'x2', 'x3']

def filter_excel(file_path, output_path, valid_names):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # 获取活动工作表

    # 遍历每一行，从第二行开始（假设第一行为表头）
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        name_cell = row[0]  # 假设姓名在第一列（A列）
        if name_cell.value not in valid_names:
            rows_to_delete.append(name_cell.row)  # 记录需要删除的行号

    # 删除记录的行（从后往前删除，避免索引错乱）
    for row_idx in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_idx)

    # 保存修改后的文件
    wb.save(output_path)
    wb.close()

# 处理两个文件
filter_excel(file1, output_file1, valid_names)
filter_excel(file2, output_file2, valid_names)

print("过滤完成，结果已保存到新的 Excel 文件中。")