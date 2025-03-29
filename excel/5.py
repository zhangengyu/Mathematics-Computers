import openpyxl
import os

# 文件夹路径
folder_path = r"C:\Users\詹耿羽\Desktop\新建文件夹"

# 检查规则：姓名为 x1 时，钱必须为 12
target_name = "x1"
target_money = 12

def check_excel(file_path, target_name, target_money):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # 获取活动工作表

    # 遍历每一行，从第二行开始（假设第一行为表头）
    invalid_rows = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        name_cell = row[0]  # 假设姓名在第一列（A列）
        money_cell = row[1]  # 假设钱在第二列（B列）
        if name_cell == target_name and money_cell != target_money:
            invalid_rows.append(row)  # 记录不符合标准的行

    wb.close()
    return invalid_rows

# 遍历文件夹中的所有 Excel 文件
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):  # 只处理 .xlsx 文件
        file_path = os.path.join(folder_path, file_name)
        invalid_rows = check_excel(file_path, target_name, target_money)

        # 输出结果
        if invalid_rows:
            print(f"文件 {file_name} 中不符合标准的行：")
            for row in invalid_rows:
                print(row)
        else:
            print(f"文件 {file_name} 中所有数据均符合标准。")