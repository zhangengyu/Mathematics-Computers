import openpyxl
import os

# 文件夹路径
folder_path = r"C:\Users\詹耿羽\Desktop\新建文件夹"

# 检查规则：姓名为 x1
target_name = "x1"

def calculate_total_money(file_path, target_name):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # 获取活动工作表

    # 初始化总金额
    total_money = 0

    # 遍历每一行，从第二行开始（假设第一行为表头）
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        name_cell = row[0]  # 假设姓名在第一列（A列）
        money_cell = row[1]  # 假设钱在第二列（B列）
        if name_cell == target_name:
            total_money += money_cell  # 累加金额

    wb.close()
    return total_money

# 遍历文件夹中的所有 Excel 文件
total_money_all_files = 0
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):  # 只处理 .xlsx 文件
        file_path = os.path.join(folder_path, file_name)
        total_money = calculate_total_money(file_path, target_name)
        total_money_all_files += total_money
        print(f"文件 {file_name} 中姓名为 {target_name} 的总金额为：{total_money}")

# 输出所有文件的总金额
print(f"所有文件中姓名为 {target_name} 的总金额为：{total_money_all_files}")